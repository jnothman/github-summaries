#!/usr/bin/env python
# coding: utf-8

import argparse
import re

import requests
import pandas as pd

# TODO: More sophisticated approach to review state precedence than using max?
# TODO: Use GitHub API v4 (GraphQL) to reduce download:
"""
# Type queries into this side of the screen, and you will
# see intelligent typeaheads aware of the current GraphQL type schema,
# live syntax, and validation errors highlighted within the text.

# We'll get you started with a simple query showing your username!
query {
  repository(owner: "scikit-learn", name: "scikit-learn") {
    collaborators {
      edges {
        node {
          login
        }
      }
    }
    pullRequests(states: [OPEN], orderBy: {field:UPDATED_AT, direction: DESC}, first: 100) {
      edges {
        node {
          number,
          author {
            login
          },
          assignees(first: 100) {
            edges {
              node{
                login
              }
            }
          },
          # reviewRequests(first: 100) {
          #  edges {
          #    node{
          #      requestedReviewer {
          #        login
          #      }
          #    }
          #  }
          #}
          reviews(first:100) {
            edges{
              node {
                author {
                  login
                },
                state
              }
            }
          }

        }
      }
    }
  }
}
"""


review_state_dtype = pd.CategoricalDtype(
    ["REQUESTED", "COMMENTED", "REQUEST_CHANGES", "APPROVED", "AUTHOR"], ordered=True
)
emoji_map = {
    "REQUESTED": "?",
    "COMMENTED": "💬",
    "REQUEST_CHANGES": "❌",
    "APPROVED": "✅",
    "AUTHOR": "✏️",
}


def get_prs_with_reviews(repo, auth_token, per_page=100):
    headers = {"Authorization": "token " + auth_token}
    resp = requests.get(
        f"https://api.github.com/repos/{repo}/pulls?sort=updated&direction=desc&state=open&per_page={per_page}",
        headers=headers,
    )
    for pr in resp.json():
        pr["reviews"] = requests.get(
            pr["url"] + "/reviews?per_page=100", headers=headers
        ).json()
        yield pr


def get_collaborators(repo, auth_token):
    headers = {"Authorization": "token " + auth_token}
    return requests.get(
        f"https://api.github.com/repos/{repo}/collaborators?per_page=100",
        headers=headers,
    ).json()


def make_reviewer_matrix(pr_list, names):
    review_df = pd.DataFrame(
        [
            {"url": pr["url"], review["user"]["login"]: review["state"]}
            for pr in pr_list
            for review in pr["reviews"]
        ]
        + [{pr["user"]["login"]: "AUTHOR", "url": pr["url"]} for pr in pr_list]
        + [
            {req["login"]: "REQUESTED"}
            for pr in pr_list
            for req in pr["requested_reviewers"]
        ]
    )

    review_df = (
        review_df.set_index("url").astype(review_state_dtype).groupby(level=0).max()
    )
    matrix = (
        pd.DataFrame(pr_list)
        .set_index("url")
        .join(review_df.apply(lambda s: s.map(emoji_map), axis=0))
    )
    matrix.index = matrix.index.map(
        lambda s: "#".join(
            re.search("https://api.github.com/repos/(.*)/pulls/([0-9]+)", s).groups()
        )
    )
    names = [name for name in names if name in matrix]
    core_matrix = matrix[names].sort_index(axis=1)

    summary = (
        core_matrix.transpose()
        .unstack()
        .dropna()
        .to_frame()
        .reset_index()
        .groupby(by=["url", 0])
        .size()
        .unstack()
    )
    summary = summary.sum(axis=1).to_frame().join(summary).rename(columns={0: "⅀"})

    # TODO: label assignee
    core_matrix = (
        matrix[["title", "updated_at"]]
        .join(summary)
        .join(core_matrix)
        .fillna("")
        .loc[summary.sort_values(["✅", "⅀"], ascending=False).index]
    )
    return core_matrix


def matrix_to_excel(path, core_matrix, names):
    from openpyxl import styles
    from openpyxl.utils import get_column_letter

    names = [name for name in names if name in core_matrix]

    writer = pd.ExcelWriter(path, engine="openpyxl")
    core_matrix.to_excel(writer)
    ws = writer.sheets["Sheet1"]
    rotated = styles.Alignment(textRotation=45, horizontal="center", vertical="bottom")
    bottom = styles.Alignment(vertical="bottom")
    for row in ws["A1:ZZ1"]:
        for cell in row:
            if cell.value in names:
                cell.alignment = rotated
                ws.column_dimensions[get_column_letter(cell.column)].width = 2.6
            elif cell.value == "title":
                cell.alignment = bottom
                ws.column_dimensions[get_column_letter(cell.column)].width = 40
            else:
                cell.alignment = bottom

    for row in ws["A1:A1000"]:
        for cell in row:
            if not cell.value:
                continue
            cell.hyperlink = "https://github.com/" + "/pull/".join(
                cell.value.split("#")
            )
            cell.alignment = styles.Alignment(horizontal="right")
            cell.font = styles.Font(color="0000ff")

    ws.freeze_panes = "C2"
    writer.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-r", "--repo", action="append")
    ap.add_argument("gh_token")
    ap.add_argument("out_xlsx")
    args = ap.parse_args()

    pr_list = []
    collaborators = []
    for repo in args.repo:
        pr_list.extend(get_prs_with_reviews(repo, auth_token=args.gh_token))
        collaborators.extend(get_collaborators(repo, auth_token=args.gh_token))
    names = {user["login"] for user in collaborators}
    matrix = make_reviewer_matrix(pr_list, names)
    matrix_to_excel(args.out_xlsx, matrix, names)


if __name__ == '__main__':
    main()
